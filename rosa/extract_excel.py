"""
Extractor de archivos Excel (.xlsx, .xls)
Lee y extrae contenido de hojas de cálculo
"""

import sys
import os

def extract_xlsx(file_path, output_file):
    """Extrae contenido de archivo Excel"""
    try:
        import openpyxl
        
        wb = openpyxl.load_workbook(file_path)
        
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(f"DOCUMENTO: {os.path.basename(file_path)}\n")
            f.write("="*80 + "\n\n")
            
            for sheet_name in wb.sheetnames:
                f.write(f"\n{'='*80}\n")
                f.write(f"HOJA: {sheet_name}\n")
                f.write(f"{'='*80}\n\n")
                
                ws = wb[sheet_name]
                
                for row in ws.iter_rows(values_only=True):
                    if any(cell is not None for cell in row):
                        row_str = '\t'.join(str(cell) if cell is not None else '' for cell in row)
                        f.write(row_str + '\n')
            
            f.write("\n" + "="*80 + "\n")
            f.write("Documento procesado correctamente\n")
        
        print(f"✓ Contenido extraído a: {output_file}")
        return True
        
    except ImportError:
        print("ERROR: La librería 'openpyxl' no está instalada.")
        print("Instálala con: pip install openpyxl")
        return False
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Uso: python extract_excel.py <archivo.xlsx> <salida.txt>")
        sys.exit(1)
    
    file_path = sys.argv[1]
    output_file = sys.argv[2]
    
    if not os.path.exists(file_path):
        print(f"❌ El archivo no existe: {file_path}")
        sys.exit(1)
    
    extract_xlsx(file_path, output_file)
